# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import re
import unicodedata
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .additional_options import ADDITIONAL_OPTIONS, AdditionalOption
from .models import Questionnaire


class ExcelGenerationError(RuntimeError):
    pass


IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
APP_ROOT = Path(__file__).resolve().parent.parent
LOCAL_TEMPLATES = APP_ROOT / "templates"
PREVIOUS_CP_TEMPLATES = (
    Path(r"C:\Users\Drugalev\Documents\Codex\2026-05-21\senior-python-b2b-1-excel-2")
    / "elevator_cp_generator"
    / "templates"
)
EXCEL_IMAGE_OPTION_DIRS = {
    "finish": LOCAL_TEMPLATES / "Walls_photo",
    "signal_steel_finish": LOCAL_TEMPLATES / "Walls_photo",
    "ceiling_steel_finish": LOCAL_TEMPLATES / "Walls_photo",
    "floor_finish": LOCAL_TEMPLATES / "Floor_photo",
    "ceiling_type": LOCAL_TEMPLATES / "Ceiling_photo",
    "handrail_type": LOCAL_TEMPLATES / "Handrails_photo",
    "mirror": LOCAL_TEMPLATES / "Mirrors_photo",
    "cop_type": LOCAL_TEMPLATES / "COPHOP_photo",
    "lop_type": LOCAL_TEMPLATES / "LOP_photo",
}
FALLBACK_EXCEL_IMAGE_OPTION_DIRS = {
    "finish": PREVIOUS_CP_TEMPLATES / "Walls_photo",
    "signal_steel_finish": PREVIOUS_CP_TEMPLATES / "Walls_photo",
    "ceiling_steel_finish": PREVIOUS_CP_TEMPLATES / "Walls_photo",
    "floor_finish": PREVIOUS_CP_TEMPLATES / "Floor_photo",
    "ceiling_type": PREVIOUS_CP_TEMPLATES / "Ceiling_photo",
    "handrail_type": PREVIOUS_CP_TEMPLATES / "Handrails_photo",
    "mirror": PREVIOUS_CP_TEMPLATES / "Mirrors_photo",
    "cop_type": PREVIOUS_CP_TEMPLATES / "COPHOP_photo",
    "lop_type": PREVIOUS_CP_TEMPLATES / "LOP_photo",
}
EXCEL_IMAGE_OPTION_PREFIX_FILTERS = {
    "signal_steel_finish": ("EX-HS", "EX-MS"),
    "ceiling_steel_finish": ("EX-HS", "EX-MS", "EX-YS"),
}
EXCEL_EXCLUDED_IMAGE_OPTION_ARTICLES = {
    "cop_type": {"EX-HX99", "IC-CARD", "IC CARD"},
}
EXCEL_MATERIAL_SUMMARY_FIELDS = [
    ("side_wall_finish", "Боковые стены", "finish"),
    ("rear_wall_finish", "Задняя стена", "finish"),
    ("front_wall_finish", "Передняя стена", "finish"),
    ("floor_finish", "Пол", "floor_finish"),
    ("handrail_finish", "Материал поручня", "signal_steel_finish"),
    ("ceiling_finish", "Материал потолка", "ceiling_steel_finish"),
    ("skirting_finish", "Плинтус", "finish"),
    ("cabin_door_finish", "Двери кабины", "finish"),
    ("main_floor_landing_door_finish", "Двери на основном посадочном этаже", "finish"),
    ("other_floors_landing_door_finish", "Двери на остальных этажах", "finish"),
    ("cop_finish", "Материал панели", "signal_steel_finish"),
    ("main_floor_lop_finish", "Материал поста на основном посадочном этаже", "signal_steel_finish"),
    ("other_floors_lop_finish", "Материал постов", "signal_steel_finish"),
]
EXCEL_EQUIPMENT_SUMMARY_FIELDS = [
    ("handrail_type", "Поручень", "handrail_type"),
    ("ceiling_type", "Потолок", "ceiling_type"),
    ("mirror", "Зеркало", "mirror"),
    ("cop_type", "Панель управления", "cop_type"),
    ("main_floor_lop_type", "Пост вызова на основном посадочном этаже", "lop_type"),
    ("other_floors_lop_type", "Посты вызова", "lop_type"),
]
UNSELECTED_EXCEL_VALUES = {"", "Другое", "Другое...", "Choose an option"}
PAIRED_EXCEL_FINISH_FIELDS = {
    "handrail_type": "handrail_finish",
    "ceiling_type": "ceiling_finish",
    "cop_type": "cop_finish",
    "main_floor_lop_type": "main_floor_lop_finish",
    "other_floors_lop_type": "other_floors_lop_finish",
}
EXCEL_FINISH_VALUE_FIELDS = {
    "side_wall_finish",
    "rear_wall_finish",
    "front_wall_finish",
    "floor_finish",
    "handrail_finish",
    "ceiling_finish",
    "skirting_finish",
    "cabin_door_finish",
    "main_floor_landing_door_finish",
    "other_floors_landing_door_finish",
    "cop_finish",
    "main_floor_lop_finish",
    "other_floors_lop_finish",
}
EXCEL_ALLOWED_TEXT_FINISH_VALUES = {
    "нет",
    "под отделку",
}
EXCEL_ARTICLE_RE = re.compile(r"\b[A-Z]{1,3}-[A-Z]{0,4}\d+[A-Z]*\b", re.IGNORECASE)
EMU_PER_PIXEL = 9525
VISUAL_SUMMARY_BODY_FONT_SIZE = 12
VISUAL_SUMMARY_TITLE_FONT_SIZE = 14
VISUAL_SUMMARY_PROJECT_FONT_SIZE = 12
VISUAL_SUMMARY_GROUP_FONT_SIZE = 12
VISUAL_SUMMARY_SECTION_FONT_SIZE = 12
QUESTIONNAIRE_RUSSIAN_COLUMN_WIDTH = 52.44140625
QUESTIONNAIRE_CHINESE_COLUMN_WIDTH = 31.33203125
QUESTIONNAIRE_GROUP_COLUMN_WIDTH = 50
QUESTIONNAIRE_SINGLE_LINE_ROW_HEIGHT = 14.4
QUESTIONNAIRE_WRAPPED_ROW_HEIGHT = 28.8
VISUAL_SUMMARY_TITLE_ROW_HEIGHT = 25.95
VISUAL_SUMMARY_HEADER_ROW_HEIGHT = 22.05
VISUAL_SUMMARY_SECTION_ROW_HEIGHT = 19.95
VISUAL_SUMMARY_MATERIAL_ROW_HEIGHT = 72.0
VISUAL_SUMMARY_EQUIPMENT_ROW_HEIGHT = 118.05
QUESTIONNAIRE_PROJECT_ROW_HEIGHT = 18.0
QUESTIONNAIRE_SECTION_ROW_HEIGHT = 15.6
QUESTIONNAIRE_FACTORY_ROW_HEIGHT = 18.0
QUESTIONNAIRE_HEADER_FILL_COLOR = Color(theme=2, tint=-0.0999786370433668)
QUESTIONNAIRE_LABEL_FILL_COLOR = Color(theme=0, tint=-0.0499893185216834)
QUESTIONNAIRE_PROJECT_FONT_COLOR = "FF002060"
QUESTIONNAIRE_FACTORY_FONT_COLOR = Color(theme=8, tint=-0.499984740745262)
QUESTIONNAIRE_PRICE_NUMBER_FORMAT = r"\¥#,##0.00;\¥\-#,##0.00"
QUESTIONNAIRE_FACTORY_ROWS = {
    "Количество контейнеров",
    "Стоимость",
}
QUESTIONNAIRE_FACTORY_ROW_ORDER = {
    "Количество контейнеров": 0,
    "Стоимость": 1,
}
QUESTIONNAIRE_SECTION_TITLES = {
    "Кабина",
    "Двери кабины",
    "Двери кабины ",
    "Этажные двери",
    "Cигнализационные устройства",
    "Сигнализационные устройства",
    "Шахта",
    "Дополнительные опции",
}


def load_mapping(path: str | Path) -> dict[str, Any]:
    with Path(path).open("r", encoding="utf-8") as file:
        return json.load(file)


def generate_questionnaire_xlsx(
    template_path: str | Path | BytesIO,
    questionnaire: Questionnaire,
    mapping_path: str | Path,
    output_path: str | Path | None = None,
    *,
    include_summary_sheet: bool = True,
) -> bytes:
    mapping = load_mapping(mapping_path)
    workbook = load_workbook(template_path)
    worksheet = _find_questionnaire_sheet(workbook, mapping.get("sheet_name"))
    _set_sheet_zoom(worksheet)
    _freeze_questionnaire_label_columns(worksheet)

    first_group_col = int(mapping.get("first_group_column", 3))
    project_rows: dict[str, int] = mapping.get("project_fields", {})
    group_rows: dict[str, int] = mapping["lift_group_fields"]
    group_count = len(questionnaire.lift_groups)
    if group_count == 0:
        raise ExcelGenerationError("Нужна хотя бы одна группа лифтов.")

    _prepare_group_columns(worksheet, first_group_col, group_count)
    additional_options_by_group, additional_options = _additional_options_by_group(questionnaire)
    if additional_options and "additional_options" in group_rows:
        group_rows = dict(group_rows)
        additional_options_row = int(group_rows["additional_options"])
        _insert_additional_option_rows(
            worksheet,
            additional_options_row,
            additional_options,
            additional_options_by_group,
            first_group_col,
        )
        _shift_group_rows_after(group_rows, additional_options_row, len(additional_options))
    _ensure_factory_rows_order(worksheet)
    _clear_group_values(worksheet, first_group_col, group_count, group_rows, project_rows)

    project_name = questionnaire.project.project_name
    if project_name and "project_name" in project_rows:
        worksheet.cell(row=int(project_rows["project_name"]), column=1).value = f"Проект: {project_name}"
        worksheet.title = _worksheet_title(project_name)

    for index, group in enumerate(questionnaire.lift_groups):
        column = first_group_col + index
        for field_name, row in group_rows.items():
            if field_name == "additional_options":
                continue
            value = _questionnaire_cell_value(group, field_name)
            if value is not None:
                worksheet.cell(row=int(row), column=column).value = value
    _fit_questionnaire_layout(worksheet, first_group_col, group_count)
    _apply_questionnaire_reference_style(worksheet, first_group_col + group_count - 1)
    if include_summary_sheet:
        _append_visual_summary(workbook, questionnaire)
    else:
        _remove_visual_summary_sheet(workbook)

    if output_path:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _find_questionnaire_sheet(workbook, configured_name: str | None) -> Worksheet:
    if configured_name:
        if configured_name not in workbook.sheetnames:
            raise ExcelGenerationError(f"Лист '{configured_name}' не найден в шаблоне.")
        return workbook[configured_name]

    for worksheet in workbook.worksheets:
        a1 = str(worksheet["A1"].value or "").lower()
        b1 = str(worksheet["B1"].value or "").lower()
        if "проект" in a1 or "项目" in b1:
            return worksheet
    return workbook.worksheets[0]


def _worksheet_title(project_name: str) -> str:
    title = str(project_name).strip()
    for character in r'[]:*?/\\':
        title = title.replace(character, " ")
    title = " ".join(title.split())
    return (title or "Опросник")[:31]


def _questionnaire_cell_value(group: Any, field_name: str) -> Any:
    value = getattr(group, field_name, None)
    if _is_unselected_excel_value(value):
        return None
    if field_name in EXCEL_FINISH_VALUE_FIELDS and not _is_selected_excel_finish_value(value):
        return None

    if field_name == "lift_name":
        return _visual_lift_name(value, getattr(group, "quantity", None))

    if field_name == "main_landing_floor":
        return _excel_landing_floor_value(value)

    finish_field = PAIRED_EXCEL_FINISH_FIELDS.get(field_name)
    if not finish_field:
        return value

    if _is_no_finish_required_excel_value(value):
        return value

    finish_value = getattr(group, finish_field, None)
    if _is_unselected_excel_value(finish_value) or not _is_selected_excel_finish_value(finish_value):
        return None

    value_text = str(value).strip()
    finish_text = str(finish_value).strip()
    if _normalize_article_text(finish_text) in _normalize_article_text(value_text):
        return value_text
    return f"{value_text}, {finish_text}"


def _is_unselected_excel_value(value: Any) -> bool:
    return value is None or str(value).strip() in UNSELECTED_EXCEL_VALUES


def _is_selected_excel_finish_value(value: Any) -> bool:
    if _is_unselected_excel_value(value):
        return False
    text = str(value).strip()
    return text.lower() in EXCEL_ALLOWED_TEXT_FINISH_VALUES or bool(EXCEL_ARTICLE_RE.search(text))


def _is_no_finish_required_excel_value(value: Any) -> bool:
    if _is_unselected_excel_value(value):
        return True
    text = str(value).strip().lower()
    return text == "нет" or text.startswith("без ")


def _excel_landing_floor_value(value: Any) -> Any:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    return int(text) if re.fullmatch(r"[+-]?\d+", text) else value


def _set_sheet_zoom(worksheet: Worksheet) -> None:
    worksheet.sheet_view.zoomScale = 80
    worksheet.sheet_view.zoomScaleNormal = 80


def _freeze_questionnaire_label_columns(worksheet: Worksheet) -> None:
    worksheet.freeze_panes = "C1"


def _fit_questionnaire_layout(worksheet: Worksheet, first_group_col: int, group_count: int) -> None:
    last_group_col = first_group_col + group_count - 1
    worksheet.column_dimensions["A"].width = QUESTIONNAIRE_RUSSIAN_COLUMN_WIDTH
    worksheet.column_dimensions["B"].width = QUESTIONNAIRE_CHINESE_COLUMN_WIDTH
    for column in range(first_group_col, last_group_col + 1):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = QUESTIONNAIRE_GROUP_COLUMN_WIDTH

    for row in range(1, worksheet.max_row + 1):
        _wrap_questionnaire_row(worksheet, row, last_group_col)
        _fit_questionnaire_row_height(worksheet, row, last_group_col)


def _wrap_questionnaire_row(worksheet: Worksheet, row: int, last_col: int) -> None:
    for column in range(1, last_col + 1):
        cell = worksheet.cell(row=row, column=column)
        if cell.value in ("", None):
            continue
        current = copy(cell.alignment) if cell.alignment else Alignment()
        cell.alignment = Alignment(
            horizontal=current.horizontal,
            vertical=current.vertical or "center",
            text_rotation=current.textRotation,
            wrap_text=True,
            shrink_to_fit=current.shrinkToFit,
            indent=current.indent,
        )


def _fit_questionnaire_row_height(worksheet: Worksheet, row: int, last_col: int) -> None:
    line_count = 1
    has_value = False
    for column in range(1, last_col + 1):
        cell = worksheet.cell(row=row, column=column)
        if cell.value in ("", None):
            continue
        has_value = True
        width = worksheet.column_dimensions[get_column_letter(column)].width or 8.43
        line_count = max(line_count, _estimated_wrapped_lines(str(cell.value), width))
    if not has_value:
        return
    worksheet.row_dimensions[row].height = (
        QUESTIONNAIRE_SINGLE_LINE_ROW_HEIGHT
        if line_count <= 1
        else QUESTIONNAIRE_WRAPPED_ROW_HEIGHT
    )


def _apply_questionnaire_reference_style(worksheet: Worksheet, last_group_col: int) -> None:
    last_col = max(worksheet.max_column, last_group_col)
    last_row = _last_meaningful_questionnaire_row(worksheet, last_col)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for row in range(1, last_row + 1):
        row_title = str(worksheet.cell(row=row, column=1).value or "").strip()
        if row == 1:
            _style_questionnaire_project_row(worksheet, row, last_col, thin_border)
        elif row_title in QUESTIONNAIRE_SECTION_TITLES:
            _style_questionnaire_section_row(worksheet, row, last_col, thin_border)
        elif row_title in QUESTIONNAIRE_FACTORY_ROWS:
            _style_questionnaire_factory_row(worksheet, row, last_col, thin_border)
        else:
            _style_questionnaire_data_row(worksheet, row, last_col, thin_border)
    _clear_questionnaire_tail_style(worksheet, last_row + 1, last_col)


def _last_meaningful_questionnaire_row(worksheet: Worksheet, last_col: int) -> int:
    for row in range(worksheet.max_row, 0, -1):
        if any(worksheet.cell(row=row, column=column).value not in ("", None) for column in range(1, last_col + 1)):
            return row
    return 1


def _style_questionnaire_project_row(worksheet: Worksheet, row: int, last_col: int, border: Border) -> None:
    worksheet.row_dimensions[row].height = QUESTIONNAIRE_PROJECT_ROW_HEIGHT
    fill = PatternFill("solid", fgColor=QUESTIONNAIRE_HEADER_FILL_COLOR)
    for column in range(1, last_col + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if column == 1:
            cell.font = Font(name="Calibri", size=14, bold=True, color=QUESTIONNAIRE_PROJECT_FONT_COLOR)
        elif column == 2:
            cell.font = Font(name="Calibri", size=11, bold=True, color=QUESTIONNAIRE_PROJECT_FONT_COLOR)
        else:
            cell.font = Font(name="Calibri", size=11)


def _style_questionnaire_section_row(worksheet: Worksheet, row: int, last_col: int, border: Border) -> None:
    worksheet.row_dimensions[row].height = QUESTIONNAIRE_SECTION_ROW_HEIGHT
    fill = PatternFill("solid", fgColor=QUESTIONNAIRE_HEADER_FILL_COLOR)
    for column in range(1, last_col + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="left" if column <= 2 else "center",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=12, bold=column <= 2)


def _style_questionnaire_factory_row(worksheet: Worksheet, row: int, last_col: int, border: Border) -> None:
    worksheet.row_dimensions[row].height = QUESTIONNAIRE_FACTORY_ROW_HEIGHT
    row_title = str(worksheet.cell(row=row, column=1).value or "").strip()
    label_fill = PatternFill("solid", fgColor=QUESTIONNAIRE_HEADER_FILL_COLOR)
    data_fill = PatternFill("solid", fgColor=Color(theme=0, tint=0.0))
    for column in range(1, last_col + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.fill = label_fill if column <= 2 else data_fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="left" if column <= 2 else "center",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=14, bold=True, color=QUESTIONNAIRE_FACTORY_FONT_COLOR)
        if column > 2:
            cell.number_format = QUESTIONNAIRE_PRICE_NUMBER_FORMAT if row_title == "Стоимость" else "General"


def _style_questionnaire_data_row(worksheet: Worksheet, row: int, last_col: int, border: Border) -> None:
    label_fill = PatternFill("solid", fgColor=QUESTIONNAIRE_LABEL_FILL_COLOR)
    empty_fill = PatternFill(fill_type=None)
    for column in range(1, last_col + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.border = border
        if column <= 2:
            cell.fill = label_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=11)
        else:
            cell.fill = empty_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            value = str(cell.value or "").strip().upper()
            cell.font = Font(name="Calibri", size=11, bold=row in {2, 3} or value == "ДА")


def _clear_questionnaire_tail_style(worksheet: Worksheet, start_row: int, last_col: int) -> None:
    empty_fill = PatternFill(fill_type=None)
    empty_border = Border()
    for row in range(start_row, worksheet.max_row + 1):
        if any(worksheet.cell(row=row, column=column).value not in ("", None) for column in range(1, last_col + 1)):
            continue
        worksheet.row_dimensions[row].height = None
        for column in range(1, last_col + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.fill = empty_fill
            cell.border = empty_border
            cell.alignment = Alignment()
            cell.font = Font(name="Calibri", size=11)


def _estimated_wrapped_lines(value: str, width: float) -> int:
    lines = 0
    usable_width = max(8, int(width * 0.9))
    for raw_line in value.splitlines() or [value]:
        text = raw_line.strip()
        if not text:
            lines += 1
            continue
        lines += max(1, int((_display_width(text) + usable_width - 1) // usable_width))
    return lines


def _display_width(value: str) -> int:
    width = 0
    for character in value:
        width += 2 if unicodedata.east_asian_width(character) in {"F", "W"} else 1
    return width


def _append_visual_summary(workbook, questionnaire: Questionnaire) -> None:
    groups_with_items = [
        (
            index,
            group,
            _visual_summary_items(group, EXCEL_MATERIAL_SUMMARY_FIELDS),
            _visual_summary_items(group, EXCEL_EQUIPMENT_SUMMARY_FIELDS),
        )
        for index, group in enumerate(questionnaire.lift_groups, start=1)
    ]
    groups_with_items = [item for item in groups_with_items if item[2] or item[3]]
    if not groups_with_items:
        return

    worksheet = _create_visual_summary_sheet(workbook)
    _setup_visual_summary_sheet(worksheet)
    row = 1
    _write_visual_summary_title(worksheet, row)
    project_name = questionnaire.project.project_name
    if project_name:
        row += 1
        _write_visual_project_title(worksheet, row, project_name)
    row += 2
    for group_index, group, material_items, equipment_items in groups_with_items:
        _write_visual_group_title(worksheet, row, group_index, group)
        row += 1
        if material_items:
            _write_visual_section_title(worksheet, row, "Материалы отделки")
            row += 1
            row = _write_material_summary_rows(worksheet, row, material_items)
        if equipment_items:
            _write_visual_section_title(worksheet, row, "Оборудование")
            row += 1
            row = _write_equipment_summary_rows(worksheet, row, equipment_items)
        row += 1


def _create_visual_summary_sheet(workbook) -> Worksheet:
    sheet_name = "Саммэри"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    return workbook.create_sheet(sheet_name)


def _remove_visual_summary_sheet(workbook) -> None:
    sheet_name = "Саммэри"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])


def _setup_visual_summary_sheet(worksheet: Worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    _set_sheet_zoom(worksheet)
    widths = {
        "A": 14,
        "B": 20,
        "C": 13,
        "D": 13,
        "E": 14,
        "F": 20,
        "G": 13,
        "H": 13,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _write_visual_summary_title(worksheet: Worksheet, row: int) -> None:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = worksheet.cell(row=row, column=1)
    cell.value = "Саммэри материалов и оборудования"
    cell.font = Font(bold=True, size=VISUAL_SUMMARY_TITLE_FONT_SIZE, color="1F2937")
    cell.fill = PatternFill("solid", fgColor="EAF2FF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[row].height = VISUAL_SUMMARY_TITLE_ROW_HEIGHT


def _write_visual_project_title(worksheet: Worksheet, row: int, project_name: str) -> None:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = worksheet.cell(row=row, column=1)
    cell.value = f"Проект: {project_name}"
    cell.font = Font(bold=True, size=VISUAL_SUMMARY_PROJECT_FONT_SIZE, color="475569")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[row].height = VISUAL_SUMMARY_HEADER_ROW_HEIGHT


def _write_visual_group_title(worksheet: Worksheet, row: int, group_index: int, group: Any) -> None:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = worksheet.cell(row=row, column=1)
    cell.value = _visual_group_title(group_index, group)
    cell.font = Font(bold=True, size=VISUAL_SUMMARY_GROUP_FONT_SIZE, color="334155")
    cell.fill = PatternFill("solid", fgColor="F8FAFC")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[row].height = VISUAL_SUMMARY_HEADER_ROW_HEIGHT


def _write_visual_section_title(worksheet: Worksheet, row: int, title: str) -> None:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = worksheet.cell(row=row, column=1)
    cell.value = title
    cell.font = Font(bold=True, size=VISUAL_SUMMARY_SECTION_FONT_SIZE, color="475569")
    cell.fill = PatternFill("solid", fgColor="EEF2F7")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[row].height = VISUAL_SUMMARY_SECTION_ROW_HEIGHT


def _write_material_summary_rows(worksheet: Worksheet, row: int, items: list[tuple[str, str, Path]]) -> int:
    for row_start in range(0, len(items), 2):
        worksheet.row_dimensions[row].height = VISUAL_SUMMARY_MATERIAL_ROW_HEIGHT
        for item_index, (label, value, image_path) in enumerate(items[row_start : row_start + 2]):
            if item_index == 0:
                image_col, text_start_col, text_end_col = 1, 2, 4
            else:
                image_col, text_start_col, text_end_col = 5, 6, 8
            _style_visual_summary_card(worksheet, row, image_col, image_col, fill_color="F8FAFC")
            _style_visual_summary_card(worksheet, row, text_start_col, text_end_col, fill_color="F8FAFC")
            worksheet.cell(row=row, column=text_start_col).value = f"{label}\n{value}"
            _add_summary_image(
                worksheet,
                row,
                image_col,
                image_col,
                image_path,
                max_width=64,
                max_height=64,
            )
        row += 1
    return row


def _write_equipment_summary_rows(worksheet: Worksheet, row: int, items: list[tuple[str, str, Path]]) -> int:
    for label, value, image_path in items:
        worksheet.row_dimensions[row].height = VISUAL_SUMMARY_EQUIPMENT_ROW_HEIGHT
        _style_visual_summary_card(worksheet, row, 1, 2, fill_color="F8FAFC")
        _style_visual_summary_card(worksheet, row, 3, 8, fill_color="F8FAFC")
        worksheet.cell(row=row, column=3).value = f"{label}\n{value}"
        _add_summary_image(
            worksheet,
            row,
            1,
            2,
            image_path,
            max_width=106,
            max_height=106,
        )
        row += 1
    return row


def _style_visual_summary_card(
    worksheet: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    fill_color: str,
) -> None:
    border = Border(
        left=Side(style="thin", color="D8DEE8"),
        right=Side(style="thin", color="D8DEE8"),
        top=Side(style="thin", color="D8DEE8"),
        bottom=Side(style="thin", color="D8DEE8"),
    )
    fill = PatternFill("solid", fgColor=fill_color)
    if end_col > start_col:
        worksheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    for column in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.border = border
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.font = Font(size=VISUAL_SUMMARY_BODY_FONT_SIZE, color="0F172A")
    worksheet.cell(row=row, column=start_col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _add_summary_image(
    worksheet: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    image_path: Path,
    max_width: int,
    max_height: int,
) -> None:
    image = ExcelImage(str(image_path))
    ratio = min(max_width / max(1, image.width), max_height / max(1, image.height))
    image.width = int(image.width * ratio)
    image.height = int(image.height * ratio)
    image.anchor = _centered_image_anchor(worksheet, row, start_col, end_col, image.width, image.height)
    worksheet.add_image(image)


def _centered_image_anchor(
    worksheet: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    image_width: int,
    image_height: int,
) -> OneCellAnchor:
    area_width = sum(_column_width_pixels(worksheet, column) for column in range(start_col, end_col + 1))
    area_height = _row_height_pixels(worksheet, row)
    offset_x = max(0, int((area_width - image_width) / 2))
    offset_y = max(0, int((area_height - image_height) / 2))
    marker = AnchorMarker(
        col=start_col - 1,
        colOff=offset_x * EMU_PER_PIXEL,
        row=row - 1,
        rowOff=offset_y * EMU_PER_PIXEL,
    )
    return OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(cx=image_width * EMU_PER_PIXEL, cy=image_height * EMU_PER_PIXEL),
    )


def _column_width_pixels(worksheet: Worksheet, column: int) -> int:
    width = worksheet.column_dimensions[get_column_letter(column)].width or 8.43
    if width < 1:
        return int(width * 12)
    return int(width * 7 + 5)


def _row_height_pixels(worksheet: Worksheet, row: int) -> int:
    height = worksheet.row_dimensions[row].height or worksheet.sheet_format.defaultRowHeight or 15
    return int(height * 96 / 72)


def _visual_group_title(group_index: int, group: Any) -> str:
    lift_name = _visual_lift_name(getattr(group, "lift_name", None), getattr(group, "quantity", None))
    section = str(getattr(group, "section", "") or "").strip()
    quantity = getattr(group, "quantity", None)
    quantity_text = f", {quantity} лифта" if isinstance(quantity, int) and quantity > 1 else ""
    if lift_name and section:
        return f"{lift_name} ({section}{quantity_text})"
    return lift_name or section or f"Группа лифтов {group_index}"


def _visual_lift_name(lift_name: Any, quantity: Any) -> str:
    text = str(lift_name or "").strip()
    if not text:
        return ""
    if not isinstance(quantity, int) or quantity <= 1 or "-" in text or "–" in text:
        return text
    match = re.match(r"^(.*?)(\d+)(\D*)$", text)
    if not match:
        return text
    prefix, start_text, suffix = match.groups()
    start = int(start_text)
    end = start + quantity - 1
    width = len(start_text) if start_text.startswith("0") else 0
    end_text = f"{end:0{width}d}" if width else str(end)
    return f"{prefix}{start_text}{suffix}-{prefix}{end_text}{suffix}"


def _visual_summary_items(group: Any, fields: list[tuple[str, str, str]]) -> list[tuple[str, str, Path]]:
    items: list[tuple[str, str, Path]] = []
    for field_name, label, option_key in fields:
        value = getattr(group, field_name, None)
        image_path = _excel_image_path_for_value(option_key, value)
        if value and image_path:
            items.append((label, str(value), image_path))
    return items


def _excel_image_path_for_value(option_key: str, value: Any) -> Path | None:
    if value in ("", None):
        return None
    normalized_value = _normalize_article_text(str(value))
    if not normalized_value:
        return None
    image_files = _excel_image_files_for_key(option_key)
    for image_path in image_files:
        article = _normalized_file_article(image_path)
        if article == normalized_value or article.replace(" ", "-") == normalized_value:
            return image_path
    for image_path in sorted(image_files, key=lambda path: len(_normalized_file_article(path)), reverse=True):
        article = _normalized_file_article(image_path)
        compact_article = article.replace(" ", "-")
        if article in normalized_value or compact_article in normalized_value:
            return image_path
    return None


def _excel_image_files_for_key(option_key: str) -> list[Path]:
    folder = _excel_image_dir_for_key(option_key)
    if not folder or not folder.exists():
        return []
    return [
        path
        for path in folder.iterdir()
        if path.is_file()
        and path.suffix.lower() in IMAGE_EXTENSIONS
        and not _is_excel_excluded_image_option(option_key, path)
        and _matches_excel_image_option_filter(option_key, path)
    ]


def _excel_image_dir_for_key(option_key: str) -> Path | None:
    folder = EXCEL_IMAGE_OPTION_DIRS.get(option_key)
    if folder and folder.exists():
        return folder
    fallback_folder = FALLBACK_EXCEL_IMAGE_OPTION_DIRS.get(option_key)
    if fallback_folder and fallback_folder.exists():
        return fallback_folder
    return folder


def _matches_excel_image_option_filter(option_key: str, path: Path) -> bool:
    prefixes = EXCEL_IMAGE_OPTION_PREFIX_FILTERS.get(option_key)
    if not prefixes:
        return True
    article = _normalized_file_article(path)
    return any(article.startswith(prefix) for prefix in prefixes)


def _is_excel_excluded_image_option(option_key: str, path: Path) -> bool:
    excluded_articles = EXCEL_EXCLUDED_IMAGE_OPTION_ARTICLES.get(option_key, set())
    if not excluded_articles:
        return False
    article = _normalized_file_article(path)
    return article in excluded_articles or article.replace(" ", "-") in excluded_articles


def _normalized_file_article(path: Path) -> str:
    return _normalize_article_text(path.stem.replace("_", " "))


def _normalize_article_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().upper()


def _prepare_group_columns(worksheet: Worksheet, first_group_col: int, group_count: int) -> None:
    for column in range(first_group_col, first_group_col + group_count):
        if column > worksheet.max_column:
            worksheet.insert_cols(column)
        if column > first_group_col + 1:
            _copy_column_style(worksheet, first_group_col, column)


def _copy_column_style(worksheet: Worksheet, source_col: int, target_col: int) -> None:
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    worksheet.column_dimensions[target_letter].width = worksheet.column_dimensions[source_letter].width
    worksheet.column_dimensions[target_letter].hidden = worksheet.column_dimensions[source_letter].hidden

    for row in range(1, worksheet.max_row + 1):
        source = worksheet.cell(row=row, column=source_col)
        target = worksheet.cell(row=row, column=target_col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def _additional_options_by_group(
    questionnaire: Questionnaire,
) -> tuple[list[set[AdditionalOption]], list[AdditionalOption]]:
    selected_by_group: list[set[AdditionalOption]] = []
    selected_union: list[AdditionalOption] = []
    for group in questionnaire.lift_groups:
        group_options = _parse_additional_options(group.additional_options)
        selected_by_group.append(group_options)
        for option in ADDITIONAL_OPTIONS:
            if option in group_options and option not in selected_union:
                selected_union.append(option)
    return selected_by_group, selected_union


def _parse_additional_options(value: str | None) -> set[AdditionalOption]:
    if not value:
        return set()
    lines = {line.strip() for line in str(value).splitlines() if line.strip()}
    return {
        option
        for option in ADDITIONAL_OPTIONS
        if option.chinese in lines or option.russian in lines
    }


def _insert_additional_option_rows(
    worksheet: Worksheet,
    section_row: int,
    additional_options: list[AdditionalOption],
    selected_by_group: list[set[AdditionalOption]],
    first_group_col: int,
) -> None:
    insert_at = section_row + 1
    worksheet.insert_rows(insert_at, amount=len(additional_options))
    source_row = insert_at + len(additional_options)
    for offset, option in enumerate(additional_options):
        row = insert_at + offset
        _copy_row_style(worksheet, source_row, row)
        worksheet.cell(row=row, column=1).value = option.russian
        worksheet.cell(row=row, column=2).value = option.chinese
        for group_index, group_options in enumerate(selected_by_group):
            if option in group_options:
                worksheet.cell(row=row, column=first_group_col + group_index).value = "ДА"


def _ensure_factory_rows_order(worksheet: Worksheet) -> None:
    rows: dict[str, int] = {}
    for row in range(1, worksheet.max_row + 1):
        title = str(worksheet.cell(row=row, column=1).value or "").strip()
        if title in QUESTIONNAIRE_FACTORY_ROW_ORDER:
            rows[title] = row
    if set(rows) != set(QUESTIONNAIRE_FACTORY_ROW_ORDER):
        return
    ordered_titles = sorted(QUESTIONNAIRE_FACTORY_ROW_ORDER, key=QUESTIONNAIRE_FACTORY_ROW_ORDER.get)
    current_titles = sorted(rows, key=rows.get)
    if current_titles == ordered_titles:
        return
    source_rows = [rows[title] for title in ordered_titles]
    target_rows = sorted(source_rows)
    row_values = {
        source_row: [worksheet.cell(row=source_row, column=column).value for column in range(1, worksheet.max_column + 1)]
        for source_row in source_rows
    }
    for title, target_row in zip(ordered_titles, target_rows):
        source_row = rows[title]
        for column, value in enumerate(row_values[source_row], start=1):
            worksheet.cell(row=target_row, column=column).value = value


def _copy_row_style(worksheet: Worksheet, source_row: int, target_row: int) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column in range(1, worksheet.max_column + 1):
        source = worksheet.cell(row=source_row, column=column)
        target = worksheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def _shift_group_rows_after(group_rows: dict[str, int], row: int, amount: int) -> None:
    for field_name, field_row in list(group_rows.items()):
        if int(field_row) > row:
            group_rows[field_name] = int(field_row) + amount


def _clear_group_values(
    worksheet: Worksheet,
    first_group_col: int,
    group_count: int,
    group_rows: dict[str, int],
    project_rows: dict[str, int],
) -> None:
    rows = set(map(int, group_rows.values())) | set(map(int, project_rows.values()))
    last_col = max(worksheet.max_column, first_group_col + group_count - 1)
    for row in rows:
        for column in range(first_group_col, last_col + 1):
            worksheet.cell(row=row, column=column).value = None
