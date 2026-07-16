# -*- coding: utf-8 -*-
from __future__ import annotations

import base64
from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl import load_workbook

from src import excel_generator
from src.excel_generator import generate_questionnaire_xlsx
from src.models import LiftGroup, ProjectInfo, Questionnaire


def _questionnaire(groups: int) -> Questionnaire:
    return Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(
                section=f"Секция {index}",
                lift_name=f"Лифт {index}",
                quantity=index,
                capacity_kg=1000 + index,
                speed_ms=1.0,
                stops=9 + index,
                doors_count=9 + index,
            )
            for index in range(1, groups + 1)
        ],
    )


def test_first_group_goes_to_column_c(template_path, mapping_path):
    content = generate_questionnaire_xlsx(template_path, _questionnaire(1), mapping_path)
    ws = load_workbook(BytesIO(content)).active
    assert ws["C2"].value == "Секция 1"
    assert ws["C5"].value == 1001
    assert ws.sheet_view.zoomScale == 80
    assert ws.freeze_panes == "C1"


def test_door_model_is_written_inside_cabin_doors_section(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(
                lift_name="Л1",
                quantity=1,
                door_model="Fermator Premium",
            )
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws["A26"].value.strip() == "Двери кабины"
    assert ws["A27"].value == "Модель дверей"
    assert ws["B27"].value == "门机型号"
    assert ws["C27"].value == "Fermator Premium"
    assert ws["A28"].value == "Тип открывания дверей"
    assert ws["A29"].value == "Облицовка"
    assert ws["A30"].value == "Этажные двери"
    assert ws["A27"].fill.fgColor.type == ws["A28"].fill.fgColor.type
    assert ws["A27"].border.bottom.style == ws["A28"].border.bottom.style


def test_door_model_defaults_to_nbsl_in_questionnaire(template_path, mapping_path):
    content = generate_questionnaire_xlsx(template_path, _questionnaire(1), mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws["C27"].value == "NBSL"


def test_questionnaire_reference_header_styles_are_applied(template_path, mapping_path):
    content = generate_questionnaire_xlsx(template_path, _questionnaire(1), mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws.row_dimensions[1].height == pytest.approx(58)
    assert ws["A1"].font.sz == 11
    assert not ws["A1"].font.bold
    assert ws["A1"].font.color.rgb == "FF002060"
    assert ws["A1"].fill.fgColor.type == "theme"
    assert ws["A1"].fill.fgColor.theme == 2

    assert ws["A2"].fill.fgColor.type == "theme"
    assert ws["A2"].fill.fgColor.theme == 0
    assert ws["A2"].alignment.horizontal == "left"
    assert ws["C2"].fill.fill_type is None
    assert ws["C2"].alignment.horizontal == "center"
    assert ws["C2"].font.bold

    assert ws.row_dimensions[13].height == pytest.approx(15.6)
    assert ws["A13"].font.sz == 12
    assert ws["A13"].font.bold
    assert ws["A13"].fill.fgColor.type == "theme"
    assert ws["A13"].fill.fgColor.theme == 2
    assert ws["C13"].font.sz == 12
    assert not ws["C13"].font.bold
    assert ws["C13"].fill.fgColor.type == "theme"
    assert ws["C13"].fill.fgColor.theme == 2


def test_second_group_goes_to_column_d(template_path, mapping_path):
    content = generate_questionnaire_xlsx(template_path, _questionnaire(2), mapping_path)
    ws = load_workbook(BytesIO(content)).active
    assert ws["D2"].value == "Секция 2"
    assert ws["D8"].value == 11


def test_three_groups_use_column_e(template_path, mapping_path):
    content = generate_questionnaire_xlsx(template_path, _questionnaire(3), mapping_path)
    ws = load_workbook(BytesIO(content)).active
    assert ws["E3"].value == "Лифт 3-Лифт 5"
    assert ws["E9"].value == 12


def test_questionnaire_lift_name_uses_range_for_multi_lift_group(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(lift_name="Л1", quantity=2),
            LiftGroup(lift_name="Л3", quantity=3),
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws["C3"].value == "Л1-Л2"
    assert ws["D3"].value == "Л3-Л5"


def test_visual_lift_name_uses_range_for_multi_lift_group():
    assert excel_generator._visual_lift_name("Л1", 3) == "Л1-Л3"
    assert excel_generator._visual_lift_name("Л01", 3) == "Л01-Л03"
    assert excel_generator._visual_lift_name("Л1-Л3", 3) == "Л1-Л3"


def test_visual_group_title_includes_capacity() -> None:
    group = LiftGroup(section="Секция 1", lift_name="Л1", quantity=2, capacity_kg=1000)

    assert excel_generator._visual_group_title(1, group) == "Л1-Л2 (Секция 1, 2 лифта, 1000 кг)"


def test_visual_summary_uses_full_control_panel_and_call_post_material_labels() -> None:
    labels_by_field = {
        field_name: label
        for field_name, label, _ in excel_generator.EXCEL_MATERIAL_SUMMARY_FIELDS
    }

    assert labels_by_field["cop_finish"] == "Материал приказной панели"
    assert (
        labels_by_field["other_floors_lop_finish"]
        == "Материал вызывных постов на остальных этажах"
    )
    equipment_labels_by_field = {
        field_name: label
        for field_name, label, _ in excel_generator.EXCEL_EQUIPMENT_SUMMARY_FIELDS
    }
    assert (
        equipment_labels_by_field["other_floors_lop_type"]
        == "Посты вызовов на остальных этажах"
    )


def test_project_name_goes_to_header_and_sheet_title(template_path, mapping_path):
    content = generate_questionnaire_xlsx(template_path, _questionnaire(1), mapping_path)
    ws = load_workbook(BytesIO(content)).active
    assert ws.title == "Тестовый проект"
    assert ws["A1"].value == "Проект: Тестовый проект"
    assert ws["B1"].value == "项目"
    assert ws["A1"].font.sz == pytest.approx(11)
    assert ws["C1"].value is None


def test_preparer_is_written_below_project_name(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(
            project_name="Тестовый проект",
            address="г. Москва, ул. Примерная, д. 1",
            prepared_by="Другалёв",
        ),
        lift_groups=[
            LiftGroup(
                lift_name="Л1",
                quantity=1,
                side_wall_finish="Шлифованная нержавеющая сталь EX-HS01",
            )
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    workbook = load_workbook(BytesIO(content))
    rich_workbook = load_workbook(BytesIO(content), rich_text=True)
    questionnaire_ws = workbook.active
    rich_questionnaire_ws = rich_workbook.active
    summary_ws = workbook["Саммэри"]

    assert questionnaire_ws["A1"].value == (
        "Проект: Тестовый проект\n"
        "Адрес проекта: г. Москва, ул. Примерная, д. 1\n"
        "Заполнено: Другалёв"
    )
    assert questionnaire_ws["B1"].value == "项目\n项目地址\n负责人"
    assert questionnaire_ws["A1"].font.sz == pytest.approx(11)
    assert questionnaire_ws["B1"].font.sz == pytest.approx(11)
    rich_header = rich_questionnaire_ws["A1"].value
    assert isinstance(rich_header, CellRichText)
    assert all(isinstance(part, TextBlock) for part in rich_header)
    assert [part.font.b for part in rich_header] == [True, False, True, False, True, False]
    assert all(part.font.sz == pytest.approx(11) for part in rich_header)
    assert questionnaire_ws.row_dimensions[1].height == pytest.approx(58)
    assert questionnaire_ws["A1"].alignment.wrap_text
    assert summary_ws["A2"].value == (
        "Проект: Тестовый проект\n"
        "Адрес проекта: г. Москва, ул. Примерная, д. 1\n"
        "Заполнено: Другалёв"
    )
    assert summary_ws.row_dimensions[2].height == pytest.approx(60)
    assert summary_ws["A2"].alignment.wrap_text


def test_main_landing_floor_is_written_as_number_when_numeric(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(section="Секция 1", lift_name="Л1", quantity=1, main_landing_floor="1"),
            LiftGroup(section="Секция 2", lift_name="Л2", quantity=1, main_landing_floor="-1"),
            LiftGroup(section="Секция 3", lift_name="Л3", quantity=1, main_landing_floor="B1"),
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws["C12"].value == 1
    assert isinstance(ws["C12"].value, int)
    assert ws["D12"].value == -1
    assert isinstance(ws["D12"].value, int)
    assert ws["E12"].value == "B1"


def test_russian_and_chinese_columns_are_not_changed(template_path, mapping_path):
    before = load_workbook(template_path).active
    factory_labels = {"Стоимость", "Количество контейнеров"}
    labels_before = [
        (before.cell(row=row, column=1).value, before.cell(row=row, column=2).value)
        for row in range(2, 54)
        if str(before.cell(row=row, column=1).value or "").strip() not in factory_labels
    ]
    content = generate_questionnaire_xlsx(template_path, _questionnaire(3), mapping_path)
    after = load_workbook(BytesIO(content)).active
    labels_after = [
        (after.cell(row=row, column=1).value, after.cell(row=row, column=2).value)
        for row in range(2, 55)
        if str(after.cell(row=row, column=1).value or "").strip()
        not in factory_labels | {"Модель дверей"}
    ]
    assert labels_after == labels_before


def test_new_column_style_is_copied(template_path, mapping_path):
    content = generate_questionnaire_xlsx(template_path, _questionnaire(3), mapping_path)
    ws = load_workbook(BytesIO(content)).active
    assert ws["E5"].fill.fill_type == ws["C5"].fill.fill_type
    assert ws["E5"].font.name == ws["C5"].font.name
    assert ws.column_dimensions["E"].width == ws.column_dimensions["C"].width


def test_questionnaire_columns_and_rows_fit_long_text(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(
                section="Секция 1",
                lift_name="Л1",
                quantity=1,
                cop_type=(
                    "EX-JC99A double, Зеркальная нержавеющая сталь EX-MS02 Gold, "
                    "шлифованная нержавеющая сталь EX-HS01"
                ),
                cop_finish="шлифованная нержавеющая сталь EX-HS01",
            )
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws.column_dimensions["A"].width == pytest.approx(52.44140625)
    assert ws.column_dimensions["B"].width == pytest.approx(31.33203125)
    assert ws.column_dimensions["C"].width == pytest.approx(50)
    assert ws["C38"].alignment.wrap_text
    assert ws.row_dimensions[2].height == pytest.approx(14.4)
    assert ws.row_dimensions[38].height == pytest.approx(28.8)


def test_unselected_finishes_and_materials_are_blank_in_questionnaire(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(
                section="Секция 1",
                lift_name="Л1",
                quantity=1,
                side_wall_finish="",
                rear_wall_finish="Окрашенная сталь",
                front_wall_finish="Ламинированная панель",
                floor_finish="Керамогранит",
                skirting_finish="Нет",
                handrail_type="EX-FS01",
                ceiling_type="EX-J135",
                cop_type="EX-AC99A",
                cabin_door_finish="Зеркальная нержавеющая сталь",
                main_floor_landing_door_finish="Окрашенная сталь",
                other_floors_landing_door_finish="Ламинированная панель",
            )
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws["C18"].value is None
    assert ws["C19"].value is None
    assert ws["C20"].value is None
    assert ws["C21"].value is None
    assert ws["C22"].value is None
    assert ws["C23"].value is None
    assert ws["C24"].value == "Нет"
    assert ws["C29"].value is None
    assert ws["C33"].value is None
    assert ws["C34"].value is None
    assert ws["C38"].value is None


def test_selected_finish_articles_are_written_to_questionnaire(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(
                section="Секция 1",
                lift_name="Л1",
                quantity=1,
                side_wall_finish="Шлифованная нержавеющая сталь EX-HS01",
                floor_finish="Под отделку",
                cabin_door_finish="Цветное стекло CG-04",
                main_floor_landing_door_finish="Покрытие под дерево E-05",
            )
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws["C18"].value == "Шлифованная нержавеющая сталь EX-HS01"
    assert ws["C21"].value == "Под отделку"
    assert ws["C29"].value == "Цветное стекло CG-04"
    assert ws["C33"].value == "Покрытие под дерево E-05"


def test_cyrillic_finish_that_fits_one_line_keeps_single_row_height(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(
                section="Секция 1",
                lift_name="Л1",
                quantity=1,
                side_wall_finish="Шлифованная нержавеющая сталь EX-HS03 Bronze",
            )
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws["C18"].value == "Шлифованная нержавеющая сталь EX-HS03 Bronze"
    assert ws.row_dimensions[18].height == pytest.approx(14.4)


def test_additional_options_are_written_to_questionnaire(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(
                section="Секция 1",
                lift_name="Л1",
                quantity=1,
                additional_options="预留视频监控接口\n手势呼梯",
            )
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    ws = load_workbook(BytesIO(content)).active

    assert ws["A51"].value == "Дополнительные опции"
    assert ws["A52"].value == "Подготовка под видеонаблюдение"
    assert ws["B52"].value == "预留视频监控接口"
    assert ws["C52"].value == "ДА"
    assert ws["A53"].value == "Gesture Call"
    assert ws["B53"].value == "手势呼梯"
    assert ws["C53"].value == "ДА"
    assert ws["A54"].value == "Доступность МГН"


def test_factory_rows_use_reference_format_and_clear_blank_tail(template_path, mapping_path):
    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(
                section="Секция 1",
                lift_name="Л1",
                quantity=1,
                additional_options="预留视频监控接口\n手势呼梯",
            ),
            LiftGroup(
                section="Секция 2",
                lift_name="Л2",
                quantity=1,
                additional_options="预留视频监控接口",
            ),
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    ws = load_workbook(BytesIO(content)).active
    rows_by_label = {
        str(ws.cell(row=row, column=1).value or "").strip(): row
        for row in range(1, ws.max_row + 1)
    }
    containers_row = rows_by_label["Количество контейнеров"]
    price_row = rows_by_label["Стоимость"]

    assert containers_row < price_row
    assert ws.row_dimensions[containers_row].height == pytest.approx(18)
    assert ws.row_dimensions[price_row].height == pytest.approx(18)
    for row in (containers_row, price_row):
        assert ws.cell(row=row, column=1).fill.fgColor.type == "theme"
        assert ws.cell(row=row, column=1).fill.fgColor.theme == 2
        assert ws.cell(row=row, column=1).font.sz == 14
        assert ws.cell(row=row, column=1).font.bold
        assert ws.cell(row=row, column=1).font.color.type == "theme"
        assert ws.cell(row=row, column=1).font.color.theme == 8
        assert ws.cell(row=row, column=1).font.color.tint == pytest.approx(-0.499984740745262)
        assert ws.cell(row=row, column=3).fill.fgColor.type == "theme"
        assert ws.cell(row=row, column=3).fill.fgColor.theme == 0
        assert ws.cell(row=row, column=3).font.sz == 14
        assert ws.cell(row=row, column=3).font.bold
        assert ws.cell(row=row, column=3).font.color.type == "theme"
        assert ws.cell(row=row, column=3).font.color.theme == 8
        assert ws.cell(row=row, column=3).border.bottom.style == "thin"
        assert ws.cell(row=row, column=3).value is None
    assert ws.cell(row=containers_row, column=3).number_format == "General"
    assert "¥" in ws.cell(row=price_row, column=3).number_format

    blank_tail_row = price_row + 1
    assert all(ws.cell(row=blank_tail_row, column=column).value is None for column in range(1, 5))
    assert ws.row_dimensions[blank_tail_row].height is None
    assert ws.cell(row=blank_tail_row, column=1).fill.fill_type is None
    assert ws.cell(row=blank_tail_row, column=1).border.bottom is None
    assert ws.cell(row=blank_tail_row, column=3).fill.fill_type is None
    assert ws.cell(row=blank_tail_row, column=3).border.bottom is None


def test_visual_summary_is_added_below_questionnaire(template_path, mapping_path, monkeypatch, tmp_path):
    material_dir = tmp_path / "walls"
    equipment_dir = tmp_path / "equipment"
    material_dir.mkdir()
    equipment_dir.mkdir()
    material_image = material_dir / "EX-HS01.png"
    equipment_image = equipment_dir / "EX-AC99A.png"
    lop_image = equipment_dir / "EX-JC99A.png"
    image_bytes = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )
    material_image.write_bytes(image_bytes)
    equipment_image.write_bytes(image_bytes)
    lop_image.write_bytes(image_bytes)
    monkeypatch.setattr(
        excel_generator,
        "EXCEL_IMAGE_OPTION_DIRS",
        {
            "finish": material_dir,
            "cop_type": equipment_dir,
            "lop_type": equipment_dir,
        },
    )

    questionnaire = Questionnaire(
        project=ProjectInfo(project_name="Тестовый проект"),
        lift_groups=[
            LiftGroup(
                section="Секция 1",
                lift_name="Л1",
                quantity=1,
                capacity_kg=1000,
                side_wall_finish="Шлифованная нержавеющая сталь EX-HS01",
                main_floor_landing_door_finish="Шлифованная нержавеющая сталь EX-HS01",
                cop_type="EX-AC99A",
                main_floor_lop_type="EX-JC99A",
                other_floors_lop_type="EX-JC99A",
            )
        ],
    )

    content = generate_questionnaire_xlsx(template_path, questionnaire, mapping_path)
    workbook = load_workbook(BytesIO(content))
    questionnaire_ws = workbook.active
    ws = workbook["Саммэри"]
    questionnaire_values = [cell.value for row in questionnaire_ws.iter_rows() for cell in row if cell.value]
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]

    assert "Саммэри материалов и оборудования" not in questionnaire_values
    assert "Саммэри материалов и оборудования" in values
    assert "Проект: Тестовый проект" in values
    assert "Л1 (Секция 1, 1000 кг)" in values
    assert "Материалы отделки" in values
    assert "Оборудование" in values
    assert ws["B6"].value == "Боковые стены"
    assert ws["B7"].value == "Шлифованная нержавеющая сталь EX-HS01"
    assert ws["F6"].value == "Двери на основном посадочном этаже"
    assert ws["F7"].value == "Шлифованная нержавеющая сталь EX-HS01"
    assert ws["C9"].value == "Панель управления"
    assert ws["C10"].value == "EX-AC99A"
    assert ws["C11"].value == "Пост вызова на основном посадочном этаже"
    assert ws["C12"].value == "EX-JC99A"
    assert ws["C13"].value == "Посты вызовов на остальных этажах"
    assert ws["C14"].value == "EX-JC99A"
    assert all("ОПЭ" not in str(value) for value in values)
    assert ws["A1"].font.sz == 14
    assert ws["A1"].font.bold
    assert ws["A2"].font.sz == 14
    assert ws["A2"].font.bold
    assert ws["A4"].font.sz == 14
    assert ws["A4"].font.bold
    assert ws["A4"].font.color.rgb == "00FFFFFF"
    assert ws["A4"].fill.fgColor.rgb == "004472C4"
    assert ws["A5"].font.sz == 12
    assert ws["A5"].font.bold
    assert ws["B6"].font.sz == 12
    assert ws["B6"].font.bold
    assert not ws["B7"].font.bold
    assert ws["C9"].font.bold
    assert not ws["C10"].font.bold
    assert ws["C13"].font.bold
    assert not ws["C14"].font.bold
    assert ws["B6"].alignment.vertical == "bottom"
    assert ws["B7"].alignment.vertical == "top"
    assert ws.column_dimensions["A"].width == pytest.approx(14)
    assert ws.column_dimensions["B"].width == pytest.approx(20)
    assert ws.column_dimensions["C"].width == pytest.approx(13)
    assert ws.column_dimensions["F"].width == pytest.approx(20)
    assert ws.column_dimensions["G"].width == pytest.approx(13)
    assert ws.row_dimensions[1].height == pytest.approx(25.95)
    assert ws.row_dimensions[2].height == pytest.approx(25.95)
    assert ws.row_dimensions[4].height == pytest.approx(27)
    assert ws.row_dimensions[5].height == pytest.approx(19.95)
    assert ws.row_dimensions[6].height == pytest.approx(34.5)
    assert ws.row_dimensions[7].height == pytest.approx(37.5)
    assert ws.row_dimensions[8].height == pytest.approx(19.95)
    assert ws.row_dimensions[9].height == pytest.approx(56.25)
    assert ws.row_dimensions[10].height == pytest.approx(61.8)
    assert "B6:D6" in {str(cell_range) for cell_range in ws.merged_cells.ranges}
    assert "B7:D7" in {str(cell_range) for cell_range in ws.merged_cells.ranges}
    assert "A6:A7" in {str(cell_range) for cell_range in ws.merged_cells.ranges}
    assert "C9:H9" in {str(cell_range) for cell_range in ws.merged_cells.ranges}
    assert "C10:H10" in {str(cell_range) for cell_range in ws.merged_cells.ranges}
    assert len(ws._images) == 5
    assert questionnaire_ws.sheet_view.zoomScale == 80
    assert ws.sheet_view.zoomScale == 80
    with ZipFile(BytesIO(content)) as archive:
        drawing_xml = "\n".join(
            archive.read(name).decode("utf-8")
            for name in archive.namelist()
            if name.startswith("xl/drawings/") and name.endswith(".xml")
        )
    assert 'cx="609600" cy="609600"' in drawing_xml
    assert 'cx="1009650" cy="1009650"' in drawing_xml
    assert "<colOff>180975</colOff>" in drawing_xml
    assert "<rowOff>152400</rowOff>" in drawing_xml
    assert "<colOff>676275</colOff>" in drawing_xml
    assert "<rowOff>238125</rowOff>" in drawing_xml

def test_visual_summary_sheet_can_be_omitted(template_path, mapping_path, tmp_path):
    template_with_summary = tmp_path / "template_with_summary.xlsx"
    workbook = load_workbook(template_path)
    workbook.create_sheet("Саммэри")
    workbook.save(template_with_summary)

    content = generate_questionnaire_xlsx(
        template_with_summary,
        _questionnaire(1),
        mapping_path,
        include_summary_sheet=False,
    )

    assert "Саммэри" not in load_workbook(BytesIO(content)).sheetnames
